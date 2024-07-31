from utils import *
import time
import openpyxl
from openpyxl import Workbook
import json


product_research_url = "https://www.sellersprite.com/v2/product-research"
store_tracking_url = "https://www.sellersprite.com/v2/store-tracking"
driver = get_driver_from_local()
ALL_STORES = {}
EXCEL_FILE_PATH = "data.xlsx"


# sleep 5 seconds to see the page



# step1 open necessary tabs
# step2 wait for user to login, after login both, press a button to trigger step3
# step3 get all boxes
# <div class="content-grid-product-box">
# loop
# 悬停到box上，找到<a>标签，其中data-tips="去亚马逊查看"，点进href,找到<a>标签，id="sellerProfileTriggerId",获取href中的seller的id
# 如果找到了seller的id，go step4
# step4 进入store-tracking页面，找到内容为添加店铺的按钮，点击，找到textarea标签，输入刚刚的seller的id，点击添加
# 找到button type="submit"，点击， 找到弹出的button type="button", 里面的<span>内容为"确定"，点击
# 找到tr,找到td class="align-middle", 往下找div class='td', 往下找div，里面有4个a标签，第一个标签提取href最后的days=7，第二个提取href最后的days=15
# 点击第一个a标签，找BSR不为空的tr
# 找button, type="button", data-days="15"，找BSR不为空的tr
# bsr不为空的tr找法，tr内部第3个td(从1开始)，下面的div class='td', 找button, 能trim成数字（格式是9,019这种，要考虑逗号，或者简单点，不是'-'就行）
# 完成后回到store-tracking页面，找到<input type="checkbox" id="checkbox_t000">, 选中， 找到button, id="remove-selected-all"，点击
# 弹出的框中找到button type="button" 内容是“确定”， 点击
# 继续step3的loop
# inject 一个按钮，点击停止

def main():
    global driver, EXCEL_FILE_PATH

    # 创建一个新的 Excel 文件
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    
    # 设置第一列和第二列的列名
    sheet["A1"] = "7天"
    sheet["B1"] = "15天"

    # 保存 Excel 文件
    workbook.save(EXCEL_FILE_PATH)

    # step1 open necessary tabs
    driver.get(product_research_url)
    print(driver.title)
    driver.implicitly_wait(5.)
    

    #step2 wait for user to login, and setup the initial page
    # 等待特定的 div 出现
    WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.container.p-0.pt-3')))
    print("特定的 div 已出现，继续执行后续操作")
    wait_for_user_login()

    driver.execute_script(f"window.open(`{store_tracking_url}`, '_blank');")
    time.sleep(3)
    driver.switch_to.window(driver.window_handles[0])
    # print("user clicked")
    
    #step3 get all boxes
    WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.d-flex.flex-wrap.bg-white.pl-4.pr-4.pb-4.mb-4.position-relative')))
    container_div = driver.find_element(By.CSS_SELECTOR, 'div.d-flex.flex-wrap.bg-white.pl-4.pr-4.pb-4.mb-4.position-relative')
    content_boxes = container_div.find_elements(By.CSS_SELECTOR, 'div.content-grid-product-box')
    
    for box in content_boxes:
        process_content_boxes(box)
    # find next page
    while True:

        try:
            # 等待并找到包含 '下一页' 的 <a> 标签
            next_page_link = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//li[@class='page-item']/a[span[contains(text(), '下一页')]]"))
            )
            next_page_link.click()
            print("点击了下一页链接")
            go_to_next_page()
            
        except Exception as e:
            print(f"遍历结束")
            with open('output.json', 'w', encoding='utf-8') as json_file:
                json.dump(ALL_STORES, json_file, ensure_ascii=False, indent=4)
            break
    driver.quit()


def go_to_next_page():
    global driver
    # 等待特定的 div 出现
    WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.container.p-0.pt-3')))
    print("特定的 div 已出现，继续执行后续操作")

    time.sleep(3)
    driver.switch_to.window(driver.window_handles[0])
    
    #step3 get all boxes
    WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.d-flex.flex-wrap.bg-white.pl-4.pr-4.pb-4.mb-4.position-relative')))
    container_div = driver.find_element(By.CSS_SELECTOR, 'div.d-flex.flex-wrap.bg-white.pl-4.pr-4.pb-4.mb-4.position-relative')
    content_boxes = container_div.find_elements(By.CSS_SELECTOR, 'div.content-grid-product-box')
    
    for box in content_boxes:
        process_content_boxes(box)
    
    
    pass

def wait_for_user_login():
    global driver
    inject_continue_button_script = """
    var button = document.createElement('button');
    button.innerHTML = '请设置好筛选条件后点击此按钮继续';
    button.id = 'continueButton';
    button.style.position = 'fixed';
    button.style.top = '10px';
    button.style.right = '10px';
    button.style.zIndex = '10000';
    document.body.appendChild(button);
    
    button.addEventListener('click', function() {
        window.continueClicked = true;
    });
    """
    driver.execute_script(inject_continue_button_script)

    # 等待用户点击继续按钮
    WebDriverWait(driver, 300).until(lambda driver: driver.execute_script("return window.continueClicked === true"))
    print("用户已点击继续按钮")


def process_content_boxes(box):
    global driver
    actions = ActionChains(driver)
    seller_id = None
    # print("处理商品窗口")
    # return 
    try:
        # 悬停到box上
        actions.move_to_element(box).perform()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'a[data-tips="去亚马逊查看"]')))
    except Exception as e:
        print(f"商品窗口没有找到‘去亚马逊查看’的跳转窗口")
        return 

    try:
        # 找到a标签，其中data-tips="去亚马逊查看"
        amazon_link = box.find_element(By.CSS_SELECTOR, 'a[data-tips="去亚马逊查看"]')
        amazon_link.click()
        print("点击了去亚马逊查看的链接")
    except:
        print("点击去亚马逊查看的链接失败")
        return 

    # 切换到新打开的标签页
    driver.switch_to.window(driver.window_handles[-1])
    # 获取当前URL
    current_url = driver.current_url

    # 等待几秒看能否找到a标签id=sellerProfileTriggerId
    try:
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'sellerProfileTriggerId')))
        seller_link = driver.find_element(By.ID, 'sellerProfileTriggerId')
        url = seller_link.get_attribute('href')
        seller_id = get_seller_id(url)
        if seller_id is None:
            print(f"在{current_url}无法提取seller的id")
        else:
            print(f"seller的id是: {seller_id}")
    except:
        print(f"在{current_url}没有找到卖家信息")
    
    # 关闭当前标签页并切换回原始标签页
    driver.close()
    
    if seller_id is not None:
        # step4 进入store-tracking页面
        tracking_store(seller_id)
        pass
    time.sleep(1)
    driver.switch_to.window(driver.window_handles[0])



def switch_to_tab(pattern):
    global driver
    
    # 获取所有窗口句柄
    window_handles = driver.window_handles

    # 遍历每个窗口句柄，找到匹配的URL
    for handle in window_handles:
        driver.switch_to.window(handle)
        if driver.current_url.__contains__(pattern):
            print(f"已切换到URL为 {pattern} 的窗口")
            break
    else:
        print(f"未找到URL为 {pattern} 的窗口")
    print(driver.title)
    return


def tracking_store(store_id):
    global ALL_STORES, driver, store_tracking_url
    if store_id in ALL_STORES:
        print(f"{store_id}已经添加过了")
        return
    switch_to_tab("store_tracking")
    driver.get(store_tracking_url)
    # 点击添加店铺
    try:
        # 等待包含特定文本的按钮变得可点击
        add_store_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(., '添加店铺')]"))
        )
        add_store_button.click()
        print("点击了添加店铺的按钮")
    except Exception as e:
        print(f"点击添加店铺的按钮失败: {e}")
        return
    # 找到textarea, 输入seller_id
    try:
        time.sleep(1)  # 等待页面加载
        textarea = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//textarea[@class='form-control form-control-sm mt-3']"))
        )
        textarea.send_keys(store_id)
        print(f"输入了店铺ID: {store_id}")
    except Exception as e:
        print(f"找到textarea失败: {e}")
        return
    # 点击添加
    try:
        submit_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[@type='submit']"))
        )
        submit_button.click()
        print("点击了添加待监控的店铺按钮")
    except Exception as e:
        print(f"点击添加待监控的店铺按钮失败: {e}")
        return

    # 点击确定
    time.sleep(1)
    click_confirm_button()

    time.sleep(1)

    # 具体的数据处理
    process_store_tracking_page(store_id)

    # 回到store-tracking页面
    driver.get(store_tracking_url)
    print("导航回store-tracking页面")
    # 删除已添加的店铺
    # 找到checkbox id="checkbox_t000"
    try:
        time.sleep(2)
        find_real_checkbox_and_click()
            
    except Exception as e:
        print(f"勾选checkbox_t000失败: {e}")

    # 找到button 里面的内容包含 “批量删除“ 并点击
    try:
        delete_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(., '批量删除')]"))
        )
        time.sleep(1)
        delete_button.click()
        print("点击了批量删除按钮")
    except Exception as e:
        print(f"点击批量删除按钮失败: {e}")
        return
    
    # 找到 class 为 'swal2-actions' 的 div 元素并点击其中的确定按钮
    try:
        swal2_actions_div = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "swal2-actions"))
        )
        print("找到了 class 为 'swal2-actions' 的 div 元素")

        # 在 swal2_actions_div 中找到包含文本 '确定' 的按钮
        confirm_button = swal2_actions_div.find_element(By.XPATH, ".//button[contains(text(), '确定')]")
        
        # 点击按钮
        confirm_button.click()
        print("点击了确定按钮")
    except Exception as e:
        print(f"操作失败: {e}")
        return
    time.sleep(5)


def find_real_checkbox_and_click():
    global driver
    try:
        # 等待checkbox元素加载
        checkbox = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "checkbox_t000"))
        )
         # 使用JavaScript直接设置checkbox为选中状态
        driver.execute_script("arguments[0].checked = true;", checkbox)

        # 触发change事件，以确保状态更新
        driver.execute_script("arguments[0].dispatchEvent(new Event('change'));", checkbox)

        # 等待checkbox元素加载
        checkbox = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "checkbox_t0"))
        )
         # 使用JavaScript直接设置checkbox为选中状态
        driver.execute_script("arguments[0].checked = true;", checkbox)

        # 触发change事件，以确保状态更新
        driver.execute_script("arguments[0].dispatchEvent(new Event('change'));", checkbox)

        # 检查checkbox是否已选中，如果没有则点击选中
        if not checkbox.is_selected():
            print("Checkbox 未选中")
            checkbox.click()  # 选中checkbox
        if checkbox.is_selected():

            print("Checkbox 已选中")

    except Exception as e:
        print(f"发生错误: {e}")

    
    

def process_store_tracking_page(store_id):
    global driver,ALL_STORES
    if store_id not in ALL_STORES:
        ALL_STORES[store_id] = {}
    # 找到一个a标签，href包含days=7
    try:
        seven_days_link = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//a[contains(@href, 'days=7')]"))
        )
        seven_days_link.click()
        print("点击了7天的链接")
    except Exception as e:
        print(f"点击7天的链接失败: {e}")
        return  
    # do something
    seven_days_data = get_bsr_data(store_id)
    ALL_STORES[store_id]["7days"] = seven_days_data

    # 找到近15天上新的button
    try:
        fifteen_days_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[@type='button' and @data-days='15']"))
        )
        fifteen_days_button.click()
        print("点击了15天的按钮")
    except Exception as e:
        print(f"点击15天的按钮失败: {e}")
        return
    fiften_days_data = get_bsr_data(store_id)
    ALL_STORES[store_id]["15days"] = fiften_days_data
    append_data_to_excel(seven_days_data, fiften_days_data)


def get_bsr_data(store_id):
    global driver
    table = None
    bsr_data = []
    try:
        # 等待具有特定ID的table元素出现
        table = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, "table-condition-search"))
        )
        print("找到了ID为'table-condition-search'的表格")
    except Exception as e:
        print(f"找到ID为'table-condition-search'的表格失败")
        return bsr_data
    
    if table is None:
        return bsr_data
    # 找到table下面的tbody中的tr
    try:
        trs = table.find_elements(By.TAG_NAME, 'tr')
        for tr in trs:
            tds = tr.find_elements(By.TAG_NAME, 'td')
            if len(tds) <= 3:
                continue
            # 获取第2个子td中的商品信息
            product = tds[1].find_element(By.CLASS_NAME, 'text-primary')
            product_id = product.text.strip()
            # 获取第3个子td中的<button>元素
            button = tds[2].find_element(By.TAG_NAME, 'button')
            data = button.text.strip()  # 获取包裹在<button>里面的data
            
            # 对data进行进一步处理
            if data == '-':
                continue
            bsr_data.append(product_id)
    except:
        print("找到tbody中的tr失败")
    
    while True:
        try:
            # 等待并找到包含 '下一页' 的 <a> 标签
            next_page_link = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//a[@class='page-link' and @aria-label='Next']"))
            )
            next_page_link.click()
            print("bsr点击了下一页链接")
            nxt_data = get_bsr_next_page()
            bsr_data.extend(nxt_data)
            
        except Exception as e:
            print(f"bsr遍历结束")
            break

    return bsr_data
          

def get_bsr_next_page():
    global driver
    table = None
    bsr_data = []
    try:
        # 等待具有特定ID的table元素出现
        table = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, "table-condition-search"))
        )
        print("找到了ID为'table-condition-search'的表格")
    except Exception as e:
        print(f"找到ID为'table-condition-search'的表格失败")
        return bsr_data
    
    if table is None:
        return bsr_data
    # 找到table下面的tbody中的tr
    try:
        trs = table.find_elements(By.TAG_NAME, 'tr')
        for tr in trs:
            tds = tr.find_elements(By.TAG_NAME, 'td')
            if len(tds) <= 3:
                continue
            # 获取第2个子td中的商品信息
            product = tds[1].find_element(By.CLASS_NAME, 'text-primary')
            product_id = product.text.strip()
            # 获取第3个子td中的<button>元素
            button = tds[2].find_element(By.TAG_NAME, 'button')
            data = button.text.strip()  # 获取包裹在<button>里面的data
            
            # 对data进行进一步处理
            if data == '-':
                continue
            bsr_data.append(product_id)
    except:
        print("找到tbody中的tr失败")
    return bsr_data


def append_data_to_excel(data_7_days, data_15_days):
    # 打开已经存在的 Excel 文件
    global EXCEL_FILE_PATH
    print(f"将数据追加到 {EXCEL_FILE_PATH}\n7days : {data_7_days}\n15days : {data_15_days}")
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    sheet = workbook.active
    
    # 找到第1列的最后一行
    max_row_col1 = sheet.max_row
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value is None:
            max_row_col1 = row - 1
            break
    
    # 找到第2列的最后一行
    max_row_col2 = sheet.max_row
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=2).value is None:
            max_row_col2 = row - 1
            break
    
    # 将数据追加到第1列
    for i, value in enumerate(data_7_days, start=1):
        sheet.cell(row=max_row_col1 + i, column=1, value=value)
    
    # 将数据追加到第2列
    for i, value in enumerate(data_15_days, start=1):
        sheet.cell(row=max_row_col2 + i, column=2, value=value)
    
    # 保存 Excel 文件
    workbook.save(EXCEL_FILE_PATH)


def click_confirm_button():
    global driver
    try:
        # 等待包含特定文本的按钮变得可点击
        confirm_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(., '确定')]"))
        )
        time.sleep(1)
        confirm_button.click()
        print("点击了确定按钮")
    except Exception as e:
        print(f"点击确定按钮失败: {e}")


if __name__ == "__main__":
    main()